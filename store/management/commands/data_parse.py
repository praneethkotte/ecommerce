

import os
from pathlib import Path
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from store.models import Product

class Command(BaseCommand):
    help = 'Load data from excel'

    def handle(self, *args, **options):
        Product.objects.all().delete()
        self.stdout.write(self.style.SUCCESS('table dropped'))
        
        base_dir = Path(__file__).resolve().parent.parent.parent.parent
        book_path = os.path.join(base_dir, 'store/us_data/US_Superstore_data.xlsx')
        book = load_workbook(book_path)
        sheet = book['Orders']
        self.stdout.write(self.style.SUCCESS(sheet.title))
        max_row_num = sheet.max_row
        max_col_num = sheet.max_column
        self.stdout.write(self.style.SUCCESS(str(max_row_num)))
        self.stdout.write(self.style.SUCCESS(str(max_col_num)))

        for i in range(2, max_row_num+1): # for each row, minus the first (headers) and last (total count)
            product_name = sheet.cell(row=i, column=16).value
            if not product_name:
                continue
            sales = sheet.cell(row=i, column=18).value

            data = Product.objects.create(
                name=product_name,
                price=sales,
            )

            data.save()
        
            if (i % 100 == 0):
                self.stdout.write(self.style.SUCCESS(f'{i} records parsed'))

        self.stdout.write(self.style.SUCCESS("successfully parsed"))